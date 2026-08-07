#!/usr/bin/env python3
"""Build the Pumpini "New Outlet Setup" onboarding workbook.

Every sheet maps to a real concept in the Pumpini data model and its fields
mirror the actual backend writers (superadmin.js, stations.js, shifts.js,
prices.js, pumpService.js, seed.js). Column headers carry the exact DB column
name in parentheses so a go-live import maps 1:1.

Images (pump sample totalizer slips, tank gauge/ATG screens) cannot live in a
spreadsheet, so the sheet carries a FILE-NAME / LINK reference per pump and per
tank; the outlet supplies the photos in a shared folder and the setup step runs
OCR-confirm against them.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

FONT = "Arial"

# ── palette ───────────────────────────────────────────────────────────────
NAVY      = "1F3864"   # section titles / header band
BLUE_REQ  = "2E5496"   # required-column header
GREY_OPT  = "8497B0"   # optional-column header
GREEN_IMG = "548235"   # photo-reference column header
YELLOW    = "FFF2CC"   # cells to fill in
PHOTO_FIL = "E2EFDA"   # photo-cell tint
EX_FILL   = "EAF1FB"   # example-row tint
WHITE     = "FFFFFF"

thin = Side(style="thin", color="D0D7E5")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell(ws, coord, value, *, bold=False, size=10, color="000000", fill=None,
         italic=False, wrap=False, align="left", vcenter=True, border=False):
    c = ws[coord]
    c.value = value
    c.font = Font(name=FONT, size=size, bold=bold, italic=italic, color=color)
    c.alignment = Alignment(horizontal=align,
                            vertical="center" if vcenter else "top", wrap_text=wrap)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if border:
        c.border = BORDER
    return c

def header_row(ws, row, specs):
    """specs: list of (header, width, kind, comment). kind: 'req'|'opt'|'img'."""
    fillmap = {"req": BLUE_REQ, "opt": GREY_OPT, "img": GREEN_IMG}
    for i, (title, width, kind, note) in enumerate(specs, start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = width
        c = cell(ws, f"{col}{row}", title, bold=True, size=10, color=WHITE,
                 fill=fillmap[kind], wrap=True, align="left", border=True)
        if note:
            cm = Comment(note, "Pumpini")
            cm.width, cm.height = 280, 130
            c.comment = cm
    ws.row_dimensions[row].height = 44

def example_row(ws, row, values):
    for i, v in enumerate(values, start=1):
        col = get_column_letter(i)
        cell(ws, f"{col}{row}", v, italic=True, color="7F7F7F", fill=EX_FILL, border=True)

def input_rows(ws, first_row, specs, n=14):
    for r in range(first_row, first_row + n):
        for i, (_t, _w, kind, _c) in enumerate(specs, start=1):
            fill = PHOTO_FIL if kind == "img" else YELLOW
            cell(ws, f"{get_column_letter(i)}{r}", None, fill=fill, border=True)

def sheet_title(ws, title, subtitle, ncols):
    last = get_column_letter(ncols)
    ws.merge_cells(f"A1:{last}1")
    cell(ws, "A1", title, bold=True, size=15, color=WHITE, fill=NAVY, align="left")
    ws.row_dimensions[1].height = 26
    ws.merge_cells(f"A2:{last}2")
    cell(ws, "A2", subtitle, italic=True, size=9, color="404040", fill="F2F4F8", wrap=True)
    ws.row_dimensions[2].height = 32

def add_dv(ws, col_letter, options, first_row, last_row=200):
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(options), allow_blank=True)
    dv.errorTitle = "Choose from list"
    dv.error = "Pick a value from the list."
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")

wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════
# 0. README
# ══════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "README"
ws.sheet_view.showGridLines = False
for col, w in {"A": 3, "B": 34, "C": 78, "D": 20}.items():
    ws.column_dimensions[col].width = w

ws.merge_cells("B2:D2")
cell(ws, "B2", "Pumpini — New Outlet Setup Workbook", bold=True, size=17, color=NAVY)
ws.merge_cells("B3:D3")
cell(ws, "B3", "One workbook to stand up a brand-new outlet and all of its setup. "
               "Fill the yellow cells; work the tabs top to bottom. Photos go in a shared folder — see 'Images & photos' below.",
     italic=True, size=10, color="404040")
ws.row_dimensions[2].height = 24
ws.row_dimensions[3].height = 26

ws.merge_cells("B5:D5")
cell(ws, "B5", "Fill the tabs IN THIS ORDER — later tabs reference earlier ones",
     bold=True, size=11, color=WHITE, fill=NAVY)
ws.row_dimensions[5].height = 20

order = [
    ("1. Owner & Group", "The owner group (the billing entity) and the owner login. Everything else hangs off these."),
    ("2. Outlet",        "The station itself + all station settings (GST, address, geofence, tolerances, blind-drop, alerts)."),
    ("3. Subscription",  "The plan the outlet runs on (drives the feature ceiling / entitlement)."),
    ("4. Staff (Mgr/CCO)","Managers and back-office (CCO) logins linked to this outlet."),
    ("5. Attendants",    "The operators who run the pumps and settle shifts. Name + mobile is enough."),
    ("6. Tanks",         "Underground tanks: number, fuel, capacity, opening stock, density + opening gauge/ATG photo."),
    ("7. Pumps",         "The dispenser machines + their SAMPLE TOTALIZER-SLIP PHOTO (required for slip scanning)."),
    ("8. Nozzles",       "Each nozzle mapped to its tank + fuel (and optionally its dispenser / slip line)."),
    ("9. Fuel Prices",   "Opening per-litre price for each fuel this outlet sells."),
    ("10. Shift Timings","The 1–3 daily shift windows the outlet runs."),
    ("11. Credit Customers","Corporate / credit accounts and their go-live opening balances."),
    ("12. Credit Slip Books","Coupon booklets allocated to a credit customer: series range + a sample coupon photo."),
    ("13. Products",     "OPTIONAL — shop / lubricant products, only if the outlet uses the store module."),
    ("14. Corporate Drivers","OPTIONAL — drivers/vehicles under a credit customer (RFID/FastTag fleets)."),
]
r = 6
for name, desc in order:
    cell(ws, f"B{r}", name, bold=True, size=10, color=NAVY, border=True, fill="F2F4F8")
    cell(ws, f"C{r}", desc, size=10, wrap=True, border=True)
    ws.row_dimensions[r].height = 30
    r += 1

# ── Images & photos ───────────────────────────────────────────────────────
r += 1
cell(ws, f"B{r}", "Images & photos  (READ THIS)", bold=True, size=11, color=WHITE, fill=GREEN_IMG)
ws.merge_cells(f"C{r}:D{r}")
cell(ws, f"C{r}", "", fill=GREEN_IMG)
r += 1
img_lines = [
    "Some of the setup CANNOT be done from text alone — it needs photographs. A spreadsheet can't hold images, so the "
    "sheet holds a FILE-NAME (or a link) per photo and you supply the actual pictures alongside it.",
    "How: put every photo in ONE shared folder (Google Drive / OneDrive) and name each file after its row — e.g. "
    "'Pump-1-slip.jpg', 'Tank-1-gauge.jpg'. Type that file name (or a direct link) into the green photo column. "
    "Then share the folder link with us.",
    "Pump sample slip (green column on tab 7) — REQUIRED for slip scanning. A photo of the totalizer slip the "
    "dispenser prints. Setup runs OCR on it to read the pump SERIAL, MODEL and which line is which nozzle, and keeps "
    "it as the reference print-format for every future settlement scan. Leave Serial/Model blank if the photo is clear "
    "— OCR fills them, and the owner confirms before it's saved.",
    "Tank gauge / ATG (Pinelabs) screen (green column on tab 6) — recommended at go-live. A photo of the gauge screen "
    "showing opening stock; it's the picture behind the first wet-stock figure the outlet is measured against.",
    "Credit-coupon sample (green column on tab 12) — one photo per credit customer's booklet. It records the SLIP "
    "FORMAT the coupon parser will read (Coupon No, Date, Vehicle, litres, seal). A sample per customer/dealer, not "
    "per coupon — the live coupons are photographed per sale in the app.",
    "Nozzle meter photos are NOT a setup item — those are captured per shift during settlement, in the app.",
]
for t in img_lines:
    cell(ws, f"B{r}", "•", bold=True, color=GREEN_IMG)
    ws.merge_cells(f"C{r}:D{r}")
    cell(ws, f"C{r}", t, size=10, wrap=True, vcenter=False)
    ws.row_dimensions[r].height = 44
    r += 1

# ── Legend ────────────────────────────────────────────────────────────────
r += 1
cell(ws, f"B{r}", "Legend", bold=True, size=11, color=WHITE, fill=NAVY)
ws.merge_cells(f"C{r}:D{r}")
cell(ws, f"C{r}", "", fill=NAVY)
r += 1
legend = [
    (YELLOW,   "Yellow cell", "A value YOU fill in."),
    (BLUE_REQ, "Blue header (white text)", "REQUIRED column — the row is rejected without it."),
    (GREY_OPT, "Grey header (white text)", "Optional column — safe to leave blank; a sensible default applies."),
    (GREEN_IMG,"Green header + green cell", "A PHOTO reference — put the file name / link here; the picture goes in the shared folder."),
    (EX_FILL,  "Italic grey row", "An EXAMPLE row showing the expected format. Delete it before import."),
]
for fill, label, desc in legend:
    cell(ws, f"B{r}", "", fill=fill, border=True)
    cell(ws, f"C{r}", f"{label} — {desc}", size=10, wrap=True)
    ws.row_dimensions[r].height = 20
    r += 1

# ── House rules ───────────────────────────────────────────────────────────
r += 1
cell(ws, f"B{r}", "House rules", bold=True, size=11, color=WHITE, fill=NAVY)
ws.merge_cells(f"C{r}:D{r}")
cell(ws, f"C{r}", "", fill=NAVY)
r += 1
rules = [
    "Mobile numbers: 10 Indian digits (e.g. 9876543210). +91 is added automatically. A phone is a login ID — it must be unique.",
    "Dates: DD-MMM-YYYY (e.g. 05-Aug-2026). India is DD/MM — never MM/DD.",
    "Times: 24-hour HH:MM (e.g. 06:00, 14:00, 22:00).",
    "Fuel types must be one of: petrol, diesel, premium_petrol, cng, xp95, speed, power. Use the SAME spelling on Tanks, Nozzles and Fuel Prices.",
    "New staff/owners get the starter password 'Welcome@123' and are forced to reset on first login — leave the password blank to use it.",
    "Money is in ₹. Opening stock is in litres. Density is kg/L (e.g. petrol ≈ 0.735, diesel ≈ 0.825).",
    "This sheet COLLECTS the data + photo references; a superadmin runs the actual go-live. Nothing here writes to production by itself.",
]
for t in rules:
    cell(ws, f"B{r}", "•", bold=True, color=NAVY)
    ws.merge_cells(f"C{r}:D{r}")
    cell(ws, f"C{r}", t, size=10, wrap=True, vcenter=False)
    ws.row_dimensions[r].height = 28
    r += 1

FUELS = ["petrol", "diesel", "premium_petrol", "cng", "xp95", "speed", "power"]
LANGS = ["en", "te", "ta", "hi"]

# ══════════════════════════════════════════════════════════════════════════
# 1. OWNER & GROUP
# ══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("1. Owner & Group")
ws.sheet_view.showGridLines = False
sheet_title(ws, "1 · Owner & Owner Group",
            "The owner group is the billing entity that can hold several outlets; the owner is the top login. "
            "Create these first — the outlet is attached to them. One row each.", 6)
specs = [
    ("Owner Group name\n(owner_groups.name)", 26, "req", "The billing entity / brand that owns this outlet. Can span multiple outlets."),
    ("Group description\n(owner_groups.description)", 26, "opt", "Free text. Optional."),
    ("Billing email\n(subscriptions.billing_email)", 26, "opt", "Where invoices / billing go."),
    ("Owner name\n(users.name)", 22, "req", "The owner's full name."),
    ("Owner mobile\n(users.phone)", 18, "req", "10 digits. This is the owner's login ID — must be unique."),
    ("Owner email\n(users.email)", 26, "opt", "Optional login/contact email."),
]
header_row(ws, 4, specs)
example_row(ws, 5, ["Kamala Fuels", "Kamala group of outlets", "billing@kamala.in",
                    "Ramesh Kamala", "9876500001", "ramesh@kamala.in"])
input_rows(ws, 6, specs, n=6)
ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════
# 2. OUTLET (station + station_settings)
# ══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("2. Outlet")
ws.sheet_view.showGridLines = False
sheet_title(ws, "2 · Outlet (station + settings)",
            "The station and all of its settings on ONE row. Blue = required to create the outlet; "
            "the rest are station settings with sensible defaults if left blank.", 24)
specs = [
    ("Outlet name\n(stations.name)", 22, "req", "e.g. 'Kamala Highway'. Shown everywhere."),
    ("Oil company\n(stations.oil_company)", 14, "req", "HPCL / BPCL / IOCL / Reliance / Nayara / Shell."),
    ("Address\n(stations.address)", 30, "opt", "Street address."),
    ("City\n(stations.city)", 14, "opt", None),
    ("State\n(stations.state)", 16, "opt", None),
    ("Pincode\n(station_settings.pincode)", 10, "opt", None),
    ("GSTIN\n(gst_number / gstn)", 18, "opt", "15-char GST number. Used on invoices."),
    ("PAN\n(station_settings.pan)", 14, "opt", None),
    ("TAN\n(station_settings.tan)", 14, "opt", None),
    ("Owner WhatsApp\n(owner_whatsapp)", 16, "opt", "10 digits. Where owner alerts/reports go."),
    ("Owner email\n(station_settings.owner_email)", 22, "opt", None),
    ("Language\n(language)", 11, "opt", "Manager UI language. en / te / ta / hi. Default en."),
    ("Invoice prefix\n(invoice_prefix)", 13, "opt", "Prefix on GST invoice numbers. Default INV."),
    ("Variance threshold ₹\n(variance_threshold)", 15, "opt", "Cash short/over alert threshold. Default 50."),
    ("Latitude\n(latitude)", 12, "opt", "For geofenced attendance. Optional."),
    ("Longitude\n(longitude)", 12, "opt", "For geofenced attendance. Optional."),
    ("Geofence radius m\n(geo_fence_radius)", 13, "opt", "Default 500 m."),
    ("Geofence on?\n(geo_fence_enabled)", 12, "opt", "yes/no. Default no."),
    ("Manager blind-drop?\n(manager_blind_drop)", 14, "opt", "yes/no — mask shift sales from the manager. Default no."),
    ("Stock tol % petrol\n(stock_tol_pct_petrol)", 13, "opt", "Dip-vs-book tolerance. Default 0.75 (%)."),
    ("Stock tol % diesel\n(stock_tol_pct_diesel)", 13, "opt", "Default 0.50 (%)."),
    ("Stock tol floor L\n(stock_tol_floor_ltrs)", 13, "opt", "Minimum litres tolerance. Default 20."),
    ("Deposit alert days\n(deposit_alert_days)", 13, "opt", "Alert if no bank deposit in N days. Default 2."),
    ("Tally company\n(tally_company_name)", 20, "opt", "Company name as it appears in Tally. Optional."),
]
header_row(ws, 4, specs)
example_row(ws, 5, ["Kamala Highway", "HPCL", "NH-44, Kompally", "Hyderabad", "Telangana",
                    "500014", "36ABCDE1234F1Z5", "ABCDE1234F", "HYDK12345A", "9876500001",
                    "owner@kamala.in", "en", "KAM", 50, "17.5810", "78.4990", 500, "no",
                    "no", 0.75, 0.50, 20, 2, "Kamala Highway HPCL"])
input_rows(ws, 6, specs, n=4)
add_dv(ws, "L", LANGS, 6)
for colL in ("R", "S"):
    add_dv(ws, colL, ["yes", "no"], 6)
ws.freeze_panes = "B5"

# ══════════════════════════════════════════════════════════════════════════
# 3. SUBSCRIPTION
# ══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("3. Subscription")
ws.sheet_view.showGridLines = False
sheet_title(ws, "3 · Subscription / Plan",
            "The plan this outlet runs on. The plan drives the feature ceiling (entitlement): a plan whose name "
            "contains 'lite' → the Lite feature set, otherwise the full Pumpini set. One row.", 4)
specs = [
    ("Plan\n(station_subscriptions.plan)", 18, "req", "e.g. 'pro' or 'lite'. A name containing 'lite' unlocks the Lite feature set only."),
    ("Status\n(status)", 14, "opt", "active / suspended / cancelled. Default active."),
    ("Start date\n(start_date)", 16, "opt", "Default = today."),
    ("End date\n(end_date)", 16, "opt", "Blank = open-ended / current subscription."),
]
header_row(ws, 4, specs)
example_row(ws, 5, ["pro", "active", "05-Aug-2026", ""])
input_rows(ws, 6, specs, n=3)
add_dv(ws, "A", ["pro", "lite", "trial"], 6)
add_dv(ws, "B", ["active", "suspended", "cancelled"], 6)
ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════
# 4. STAFF (managers / CCO)
# ══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("4. Staff (Mgr-CCO)")
ws.sheet_view.showGridLines = False
sheet_title(ws, "4 · Managers & Back-office (CCO)",
            "Managers run the outlet day-to-day; CCO (Central Cash Office) users get back-office access across the "
            "owner group. Attendants go on the next tab. Leave password blank to use the 'Welcome@123' starter.", 5)
specs = [
    ("Name\n(users.name)", 24, "req", None),
    ("Mobile\n(users.phone)", 16, "req", "10 digits. Login ID — must be unique."),
    ("Email\n(users.email)", 26, "opt", None),
    ("Role\n(users.role)", 14, "req", "manager = runs this outlet · cco = back-office across the owner group."),
    ("Starter password\n(password)", 18, "opt", "Leave blank for 'Welcome@123' (forced reset on first login)."),
]
header_row(ws, 4, specs)
example_row(ws, 5, ["Suresh Manager", "9876500010", "suresh@kamala.in", "manager", ""])
input_rows(ws, 6, specs, n=10)
add_dv(ws, "D", ["manager", "cco"], 6)
ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════
# 5. ATTENDANTS
# ══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("5. Attendants")
ws.sheet_view.showGridLines = False
sheet_title(ws, "5 · Attendants (operators)",
            "The operators who run the nozzles and settle shifts. Name + mobile is all that is required; they get "
            "'Welcome@123' and reset on first login. An existing number is re-activated, not duplicated.", 3)
specs = [
    ("Name\n(users.name)", 26, "req", None),
    ("Mobile\n(users.phone)", 16, "req", "10 digits. Login ID — must be unique. Both name and mobile are required."),
    ("Language\n(users.language)", 14, "opt", "en / te / ta / hi. Default en."),
]
header_row(ws, 4, specs)
example_row(ws, 5, ["Arjun S", "9876500020", "te"])
input_rows(ws, 6, specs, n=20)
add_dv(ws, "C", LANGS, 6)
ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════
# 6. TANKS  (+ opening gauge/ATG photo)
# ══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("6. Tanks")
ws.sheet_view.showGridLines = False
sheet_title(ws, "6 · Tanks",
            "One row per underground tank. 'Opening stock' is the dip-verified stock at go-live. Tank number is "
            "referenced by nozzles on tab 8. The green column is the file name / link of the opening gauge photo.", 6)
specs = [
    ("Tank #\n(tanks.tank_number)", 10, "req", "1, 2, 3 … Unique within the outlet. Nozzles refer to this."),
    ("Fuel type\n(tanks.fuel_type)", 16, "req", "petrol / diesel / premium_petrol / cng / xp95 / speed / power."),
    ("Capacity L\n(capacity_ltrs)", 14, "req", "Nominal tank capacity in litres."),
    ("Opening stock L\n(current_stock)", 16, "req", "Dip-verified litres in the tank at go-live."),
    ("Density kg/L\n(density)", 14, "opt", "e.g. petrol ≈ 0.735, diesel ≈ 0.825. Optional but recommended."),
    ("Opening gauge/ATG photo\n(file name / link → station_artifacts)", 26, "img",
     "Photo of the gauge / ATG (Pinelabs) screen showing this tank's opening stock at go-live. "
     "Put the file name (e.g. Tank-1-gauge.jpg) or a link here; the picture goes in the shared folder. Recommended."),
]
header_row(ws, 4, specs)
example_row(ws, 5, [1, "petrol", 20000, 15000, 0.7350, "Tank-1-gauge.jpg"])
input_rows(ws, 6, specs, n=12)
add_dv(ws, "B", FUELS, 6)
ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════
# 7. PUMPS / DISPENSERS  (+ sample totalizer-slip photo)
# ══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("7. Pumps")
ws.sheet_view.showGridLines = False
sheet_title(ws, "7 · Pumps / Dispensers  (+ sample slip photo)",
            "The physical dispenser machines. A pump is a MACHINE, not a fuel — several nozzles (grades) sit on it. "
            "The green column is the SAMPLE TOTALIZER-SLIP PHOTO — required to enable slip scanning at settlement.", 5)
specs = [
    ("Pump no\n(pumps.pump_number)", 12, "req", "The number painted on the unit, e.g. 1, 2. Nozzles refer to this."),
    ("Sample totalizer-slip photo\n(file name / link → pumps.slip_artifact_id)", 30, "img",
     "REQUIRED for slip scanning. A photo of the slip this dispenser (Pinelabs / totalizer) prints. Setup OCRs it to "
     "read the SERIAL, MODEL and nozzle lines, and keeps it as the reference print-format. Put the file name "
     "(e.g. Pump-1-slip.jpg) or a link here; the picture goes in the shared folder."),
    ("Serial\n(pumps.serial)", 16, "opt", "OCR reads this from the slip photo — leave blank if the photo is clear. The owner confirms it before it saves."),
    ("Model\n(pumps.model)", 18, "opt", "OCR reads this from the slip photo too. Leave blank if the photo is clear."),
    ("Notes\n(pumps.notes)", 22, "opt", None),
]
header_row(ws, 4, specs)
example_row(ws, 5, ["1", "Pump-1-slip.jpg", "", "", "Front island — OCR fills serial/model"])
input_rows(ws, 6, specs, n=10)
ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════
# 8. NOZZLES
# ══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("8. Nozzles")
ws.sheet_view.showGridLines = False
sheet_title(ws, "8 · Nozzles",
            "One row per nozzle. Each nozzle belongs to a tank (match the Tank # from the Tanks tab) and dispenses "
            "that tank's fuel. Pump no + slip line are optional (only if totalizer-slip scanning is used).", 5)
specs = [
    ("Nozzle #\n(nozzles.nozzle_number)", 12, "req", "Label printed on the nozzle, e.g. 1, 2, 3A. Convention '1.1' = pump1.nozzle1."),
    ("Tank #\n(→ tanks.tank_number)", 10, "req", "Which tank feeds this nozzle. Must exist on the Tanks tab."),
    ("Fuel type\n(nozzles.fuel_type)", 16, "req", "Should match the tank's fuel."),
    ("Pump no\n(→ pumps.pump_number)", 14, "opt", "OPTIONAL. Which dispenser (tab 7) this nozzle sits on."),
    ("Slip line no\n(slip_nozzle_no)", 14, "opt", "OPTIONAL. The line this nozzle prints on the dispenser's totalizer slip."),
]
header_row(ws, 4, specs)
example_row(ws, 5, ["1.1", 1, "petrol", "1", 1])
input_rows(ws, 6, specs, n=16)
add_dv(ws, "C", FUELS, 6)
ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════
# 9. FUEL PRICES
# ══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("9. Fuel Prices")
ws.sheet_view.showGridLines = False
sheet_title(ws, "9 · Opening Fuel Prices",
            "The per-litre selling price of each fuel at go-live. One row per fuel type the outlet sells. "
            "Prices are updated daily in-app after go-live; this just seeds the first price.", 2)
specs = [
    ("Fuel type\n(fuel_prices.fuel_type)", 18, "req", "petrol / diesel / premium_petrol / cng / xp95 / speed / power."),
    ("Price ₹/L\n(fuel_prices.price)", 16, "req", "Opening per-litre price."),
]
header_row(ws, 4, specs)
example_row(ws, 5, ["petrol", 105.50])
input_rows(ws, 6, specs, n=8)
add_dv(ws, "A", FUELS, 6)
ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════
# 10. SHIFT TIMINGS
# ══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("10. Shift Timings")
ws.sheet_view.showGridLines = False
sheet_title(ws, "10 · Shift Timings",
            "The daily shift windows this outlet runs (1 to 3 shifts). Times are 24-hour HH:MM. "
            "A shift that crosses midnight is fine (e.g. 22:00 → 06:00).", 4)
specs = [
    ("Shift #\n(shift_number)", 10, "req", "1, 2 or 3."),
    ("Name\n(shift_definitions.name)", 20, "req", "e.g. Morning / Afternoon / Night."),
    ("Start time\n(start_time)", 14, "req", "24-hour HH:MM, e.g. 06:00."),
    ("End time\n(end_time)", 14, "req", "24-hour HH:MM, e.g. 14:00."),
]
header_row(ws, 4, specs)
example_row(ws, 5, [1, "Morning", "06:00", "14:00"])
input_rows(ws, 6, specs, n=3)
add_dv(ws, "A", ["1", "2", "3"], 6)
ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════
# 11. CREDIT CUSTOMERS
# ══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("11. Credit Customers")
ws.sheet_view.showGridLines = False
sheet_title(ws, "11 · Credit (Corporate) Customers",
            "Corporate / credit accounts that fuel on account. 'Opening balance' is the receivable outstanding at "
            "go-live (an opening invoice is raised for it). Dedupe is on mobile within the outlet.", 5)
specs = [
    ("Company name\n(corporate_accounts.company_name)", 28, "req", "The account name."),
    ("Contact mobile\n(contact_phone)", 16, "opt", "10 digits. Used to dedupe within the outlet."),
    ("Opening balance ₹\n(→ opening invoice)", 18, "opt", "Amount owed at go-live. 0 or blank if none."),
    ("Credit limit ₹\n(corporate_station_links.credit_limit)", 16, "opt", "Per-outlet credit ceiling. Optional."),
    ("GSTIN\n(corporate_accounts.gstn)", 18, "opt", "For their GST invoices. Optional."),
]
header_row(ws, 4, specs)
example_row(ws, 5, ["ABC Logistics Pvt Ltd", "9876543210", 45000, 200000, "36ABCDE1234F1Z5"])
input_rows(ws, 6, specs, n=16)
ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════
# 12. CREDIT SLIP BOOKS (coupon booklet allocation + sample coupon photo)
# ══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("12. Credit Slip Books")
ws.sheet_view.showGridLines = False
sheet_title(ws, "12 · Credit Slip Books (coupon booklets)",
            "The pre-printed coupon booklets allocated to a credit customer. One row per booklet. The coupon series "
            "range is recorded so a missing / re-used coupon can be caught; the green column holds a sample coupon "
            "photo that records the slip format. Company name must match a row on tab 11.", 8)
specs = [
    ("Company name\n(→ Credit Customers)", 26, "req", "Which credit customer this booklet is issued to. Must match a company on tab 11."),
    ("Book label\n(credit_slip_books.book_label)", 16, "opt", "Optional printed book id / name. Free text."),
    ("Series start\n(series_start)", 14, "req", "The FIRST coupon number in the book, e.g. 17701. Ranges must not overlap another book at this outlet."),
    ("Series end\n(series_end)", 14, "req", "The LAST coupon number in the book, e.g. 17800. Must be ≥ series start."),
    ("First unused #\n(opening_leaf)", 14, "opt", "Only if the book is handed over PART-USED — the first still-blank coupon. Blank = starts at series start."),
    ("Issued on\n(issued_on)", 14, "opt", "Date the book was handed over. Default = today."),
    ("Status\n(status)", 13, "opt", "active / exhausted / cancelled / lost. Default active."),
    ("Sample coupon photo\n(→ credit_slip_books.sample_artifact_id)", 28, "img",
     "One photo of a coupon from THIS booklet — the slip format the parser reads (Coupon No, Date, Vehicle, litres, "
     "seal) AND the dispute backup a bill is checked against later. Stored against the booklet as a coupon artifact. "
     "Put the file name (e.g. Coupon-ABC-sample.jpg) or a link here; the picture goes in the shared folder."),
]
header_row(ws, 4, specs)
example_row(ws, 5, ["ABC Logistics Pvt Ltd", "Book-A/26-27", 17701, 17800, "", "05-Aug-2026",
                    "active", "Coupon-ABC-sample.jpg"])
input_rows(ws, 6, specs, n=12)
add_dv(ws, "G", ["active", "exhausted", "cancelled", "lost"], 6)
ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════
# 13. PRODUCTS (optional)
# ══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("13. Products")
ws.sheet_view.showGridLines = False
sheet_title(ws, "13 · Products  (OPTIONAL — only if the store module is used)",
            "Shop / lubricant products sold at the outlet's store. Skip this tab entirely if the outlet only sells fuel.", 8)
specs = [
    ("Product name\n(products.name)", 26, "req", None),
    ("Brand\n(products.brand)", 16, "opt", None),
    ("Unit\n(products.unit)", 12, "opt", "piece / litre / kg. Default piece."),
    ("Selling price ₹\n(selling_price)", 14, "req", None),
    ("Buying price ₹\n(buying_price)", 14, "opt", None),
    ("GST %\n(gst_rate)", 10, "opt", "Default 18."),
    ("HSN code\n(hsn_code)", 14, "opt", None),
    ("Opening stock\n(current_stock)", 14, "opt", "Units in stock at go-live."),
]
header_row(ws, 4, specs)
example_row(ws, 5, ["Servo 4T 900ml", "Servo", "piece", 480, 360, 18, "27101980", 24])
input_rows(ws, 6, specs, n=14)
ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════
# 13. CORPORATE DRIVERS (optional)
# ══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("14. Corporate Drivers")
ws.sheet_view.showGridLines = False
sheet_title(ws, "14 · Corporate Drivers / Vehicles  (OPTIONAL)",
            "Drivers and vehicles under a credit customer — used by RFID / FastTag fleet fuelling. "
            "The 'Company name' must match a row on the Credit Customers tab.", 6)
specs = [
    ("Company name\n(→ Credit Customers)", 26, "req", "Must match a company on tab 11."),
    ("Driver name\n(corporate_drivers.name)", 20, "req", None),
    ("Driver mobile\n(phone)", 16, "opt", "10 digits."),
    ("Vehicle number\n(vehicle_number)", 16, "opt", "e.g. TS09AB1234."),
    ("FastTag / RFID id\n(fasttag_id)", 18, "opt", None),
    ("Per-fill limit ₹\n(per_fill_limit)", 15, "opt", "Max value per fuelling. Optional."),
]
header_row(ws, 4, specs)
example_row(ws, 5, ["ABC Logistics Pvt Ltd", "Rajan M", "9876000001", "TS09AB1234", "FT123456", 5000])
input_rows(ws, 6, specs, n=14)
ws.freeze_panes = "A5"

out = "/home/user/Pumpini/docs/New-Outlet-Setup.xlsx"
wb.save(out)
print("saved", out)
